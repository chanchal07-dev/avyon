from flask import Flask, render_template, request, jsonify
from flask_cors import CORS
from openpyxl import Workbook, load_workbook
import os
from datetime import datetime

app = Flask(__name__)
CORS(app)

# Excel file path
EXCEL_FILE = 'form_responses.xlsx'

def init_excel_file():
    """Initialize Excel file with headers if it doesn't exist"""
    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        ws = wb.active
        ws.title = "Form Responses"
        # Add headers
        headers = [
            'Name', 'Roll No', 'Hobby', 'Fantasy', 'Guilty Pleasure', 
            'Senior Date', 'Secret', 'Excitement Rating', 'Titles',
            'Timestamp'
        ]
        ws.append(headers)
        wb.save(EXCEL_FILE)

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/submit', methods=['POST'])
def submit_form():
    try:
        data = request.json
        print("Form data received:", data)
        
        # Initialize Excel file
        init_excel_file()
        
        # Load existing workbook
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active
        
        # Prepare data for Excel
        row_data = [
            data.get('name', ''),
            data.get('rollno', ''),
            data.get('hobby', ''),
            data.get('fantasy', ''),
            data.get('guilty-pleasure', ''),
            data.get('senior-date', ''),
            data.get('secret', ''),
            data.get('excitementRating', ''),
            ', '.join(data.get('titles', [])),
            data.get('timestamp', '')
        ]
        
        # Append new row
        ws.append(row_data)
        
        # Save the workbook
        wb.save(EXCEL_FILE)
        
        return jsonify({
            "status": "success", 
            "message": "Form submitted successfully and saved to Excel!"
        })
        
    except Exception as e:
        print("Error:", str(e))
        return jsonify({
            "status": "error", 
            "message": f"Error saving data: {str(e)}"
        }), 500

if __name__ == '__main__':
    app.run(debug=True)
